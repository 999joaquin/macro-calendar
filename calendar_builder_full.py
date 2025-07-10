# calendar_builder_full.py
"""
Automated macro-economic calendar
=================================
• Live data (next ~2 weeks) via TradingEconomics Developer API
• Static scheduled events (Jul 2025 – Jan 2026) auto-generated
• Outputs/updates macro_calendar.xlsx (Calendar & Glossary sheets)
---------------------------------------------------------------
Requirements:
  pip install pandas openpyxl requests python-dateutil python-dotenv
.env file:
  TE_CLIENT=d7ce0826836148f
  TE_SECRET=uk5cjdr5i1qzbzr
"""

import os, datetime as dt
from pathlib import Path
from dotenv import load_dotenv
import requests, pandas as pd
from dateutil.relativedelta import relativedelta, FR, TH
from openpyxl import load_workbook

# ------------------------------------------------------------------ CONFIG
load_dotenv()                                # developer API creds
CLIENT  = os.getenv("TE_CLIENT",  "d7ce0826836148f")
SECRET  = os.getenv("TE_SECRET",  "uk5cjdr5i1qzbzr")

START_STATIC = dt.date(2025, 7, 1)           # static horizon
END_STATIC   = dt.date(2026, 1, 31)

EXCEL_FILE = Path("macro_calendar.xlsx")

# ------------------------------------------------------------------ 1. LIVE DATA (next ~14 days)
def fetch_live_te():
    today = dt.date.today()
    horizon = today + dt.timedelta(days=14)          # 2-week look-ahead
    url = "https://api.tradingeconomics.com/calendar"
    params = {
        "d1": today.strftime("%Y-%m-%d"),
        "d2": horizon.strftime("%Y-%m-%d"),
        "c":  f"{CLIENT}:{SECRET}",
        "f":  "json"
    }
    print("→ Fetching live TradingEconomics window …")
    r = requests.get(url, params=params, timeout=30)
    r.raise_for_status()
    live = pd.DataFrame(r.json())
    if live.empty:
        print("⚠️ Live feed empty — continuing with static data only.")
        return pd.DataFrame()
    keep = {
        "Date":"Date", "Country":"Country", "Event":"Event",
        "Actual":"Actual", "Previous":"Previous",
        "Forecast":"Forecast", "Consensus":"Forecast",
        "Importance":"Impact"
    }
    live = live[[c for c in keep if c in live.columns]].rename(columns=keep)
    live["Date"] = pd.to_datetime(live["Date"]).dt.date
    live["Source"] = f"TE_live_{today}"
    return live

# ------------------------------------------------------------------ 2. STATIC SCHEDULED EVENTS
def monthly_on(weekday_rule, base_day):
    """Returns function to find e.g. 1st Friday (NFP) or nearest Thursday."""
    def date_in_month(d):
        return (d + relativedelta(day=base_day, weekday=weekday_rule))
    return date_in_month

def generate_static():
    rows = []
    cur = START_STATIC
    nfp_rule   = monthly_on(FR(1), 1)          # 1st Friday
    cpi_rule   = monthly_on(TH(0), 10)         # ~10th (nearest Thu)
    ppi_rule   = monthly_on(TH(0), 12)         # ~12th
    retail_rule= monthly_on(TH(0), 14)         # ~14th
    gdp_rule   = monthly_on(TH(0), 28)         # ~28th
    
    while cur <= END_STATIC:
        # US recurring
        rows += [
            (nfp_rule(cur),   "United States", "Non-Farm Payrolls",             3),
            (cpi_rule(cur),   "United States", "Consumer Price Index YoY",      2),
            (ppi_rule(cur),   "United States", "Producer Price Index YoY",      2),
            (retail_rule(cur),"United States", "Retail Sales MoM",              2),
            (gdp_rule(cur),   "United States", "GDP Advance Estimate QoQ",      3),
        ]
        cur += relativedelta(months=1)
    
    # FOMC & ECB official dates
    fomc = ["2025-07-30","2025-09-17","2025-11-05","2025-12-17","2026-01-28"]
    ecb  = ["2025-07-17","2025-09-11","2025-10-23","2025-12-04","2026-01-22"]
    for d in fomc:
        rows.append((dt.date.fromisoformat(d), "United States", "FOMC Meeting & Rate Decision", 3))
    for d in ecb:
        rows.append((dt.date.fromisoformat(d), "Euro Area",     "ECB Interest Rate Decision",  3))
    
    # Singapore quarterly GDP
    for d in ["2025-07-12","2025-10-11","2026-01-10"]:
        rows.append((dt.date.fromisoformat(d), "Singapore", "GDP Advance Estimate QoQ", 2))
    
    df = pd.DataFrame(rows, columns=["Date","Country","Event","Impact"])
    df["Forecast"]=df["Previous"]=df["Actual"]=""
    df["Source"]="Static_Schedule"
    return df

# ------------------------------------------------------------------ 3. MERGE & WRITE EXCEL
def write_excel(live_df, static_df):
    combined = pd.concat([live_df, static_df], ignore_index=True)\
                 .drop_duplicates(subset=["Date","Country","Event"])\
                 .sort_values("Date")
    
    mode = "a" if EXCEL_FILE.exists() else "w"
    print(f"→ Writing to {EXCEL_FILE} ({'update' if mode=='a' else 'create'}) …")
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as xls:
        combined.to_excel(xls, sheet_name="Calendar", index=False)

        # --- Glossary ---
        if mode == "a":
            wb = load_workbook(EXCEL_FILE)
            old = (pd.read_excel(EXCEL_FILE, sheet_name="Glossary")
                   if "Glossary" in wb.sheetnames else pd.DataFrame())
        else:
            old = pd.DataFrame()
        new_gloss = combined[["Event"]].drop_duplicates().assign(
                        Purpose="", Frequency="")
        gloss = (pd.concat([old, new_gloss])
                   .drop_duplicates(subset=["Event"])
                   .reset_index(drop=True))
        gloss.to_excel(xls, sheet_name="Glossary", index=False)
    print(f"✅ Done – {len(combined)} total events now in calendar.")

# ------------------------------------------------------------------ MAIN
if __name__ == "__main__":
    live = fetch_live_te()
    static = generate_static()
    write_excel(live, static)
